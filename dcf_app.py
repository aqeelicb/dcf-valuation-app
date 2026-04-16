import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.styles import Font, Alignment

# ---------------- CONFIG ----------------
st.set_page_config(page_title="DCF Terminal", layout="wide")

st.title("DCF Valuation Terminal")
st.caption("Professional Equity Valuation Model")

# ---------------- SIDEBAR ----------------
st.sidebar.image("logo.png", width=140)
st.sidebar.markdown("### Model Inputs")

company_name = st.sidebar.text_input("Company Name", key="company_name")
start_year = st.sidebar.number_input("Start Year", value=2021, key="start_year")

net_debt = st.sidebar.number_input("Net Debt", value=0.0, key="net_debt")
shares = st.sidebar.number_input("Shares Outstanding", value=1000.0, key="shares")
market_price = st.sidebar.number_input("Market Price", value=0.0, key="market_price")

wacc = st.sidebar.number_input("WACC (%)", value=15.0, key="wacc") / 100
terminal_growth = st.sidebar.number_input("Terminal Growth (%)", value=4.0) / 100
tax_rate = st.sidebar.number_input("Tax Rate (%)", value=29.0) / 100
years = int(st.sidebar.number_input("Forecast Years", value=5))

# ---------------- TABS ----------------
tab1, tab2, tab3 = st.tabs(["Data Input", "Valuation", "Sensitivity"])

# ---------------- DATA ----------------
revenues = [0.0]*5
ebit = [0.0]*5
capex = [0.0]*5
wc = [0.0]*5
da = [0.0]*5

# ---------------- TAB 1 ----------------
with tab1:
    st.subheader("Historical Financials")

    years_hist = [start_year+i for i in range(5)]

    for i in range(5):
        c1, c2, c3, c4, c5 = st.columns(5)

        revenues[i] = c1.number_input(f"Revenue {years_hist[i]}", key=f"rev{i}")
        ebit[i] = c2.number_input(f"EBIT {years_hist[i]}", key=f"ebit{i}")
        da[i] = c3.number_input(f"D&A {years_hist[i]}", key=f"da{i}")
        capex[i] = c4.number_input(f"Capex {years_hist[i]}", key=f"cap{i}")
        wc[i] = c5.number_input(f"WC Change {years_hist[i]}", key=f"wc{i}")

    
# ---------------- ASSUMPTIONS ----------------
growth_rates = [(revenues[i]-revenues[i-1])/revenues[i-1] for i in range(1,5) if revenues[i-1]!=0]
growth = np.mean(growth_rates) if growth_rates else 0

margin = np.mean([ebit[i]/revenues[i] for i in range(5) if revenues[i]!=0])
capex_pct = np.mean([capex[i]/revenues[i] for i in range(5) if revenues[i]!=0])
da_pct = np.mean([da[i]/revenues[i] for i in range(5) if revenues[i]!=0])
wc_pct = np.mean([(wc[i]/(revenues[i]-revenues[i-1])) for i in range(1,5) if revenues[i]-revenues[i-1]!=0])

# ---------------- TAB 2 ----------------
with tab2:
    st.subheader("Valuation Assumptions")

    col1, col2 = st.columns(2)

    with col1:
        growth_rate = st.slider("Growth (%)", 0.0, 20.0, float(growth*100))/100
        ebit_margin = st.slider("EBIT Margin (%)", 0.0, 60.0, float(margin*100))/100
        da_pct_user = st.slider("D&A (% Revenue)", 0.0, 10.0, float(da_pct*100))/100

    with col2:
        capex_pct_user = st.slider("Capex (% Revenue)", 0.0, 20.0, float(capex_pct*100))/100
        wc_pct_user = st.slider("WC (% ΔRevenue)", -10.0, 20.0, float(wc_pct*100))/100

    rev = revenues[-1]
    per_share = None

    if rev > 0:
        fcf = []
        prev = rev

        for t in range(years):
            rev *= (1 + growth_rate)
            delta = rev - prev

            nopat = rev * ebit_margin * (1 - tax_rate)
            da_val = rev * da_pct_user
            capex_val = rev * capex_pct_user
            wc_val = delta * wc_pct_user

            fcf.append(nopat + da_val - capex_val - wc_val)
            prev = rev

        discounted = [fcf[i]/((1+wacc)**(i+1)) for i in range(len(fcf))]
        tv = fcf[-1]*(1+terminal_growth)/(wacc-terminal_growth)
        tv_discounted = tv/((1+wacc)**years)

        ev = sum(discounted) + tv_discounted
        equity = ev - net_debt
        per_share = equity / shares
        mos = (per_share - market_price)/market_price if market_price else 0

        c1, c2, c3 = st.columns(3)
        c1.metric("Fair Value", round(per_share,2))
        c2.metric("Market Price", market_price)
        c3.metric("Margin of Safety", f"{round(mos*100,2)}%")

# ---------------- TAB 3 ----------------
with tab3:
    st.subheader("Sensitivity Analysis")

    # -------- SAFE VARIABLES --------
    try:
        g_base = growth_rate
        em_base = ebit_margin
        da_base = da_pct_user
        cap_base = capex_pct_user
        wc_base = wc_pct_user
    except:
        g_base = growth
        em_base = margin
        da_base = da_pct
        cap_base = capex_pct
        wc_base = wc_pct

    if revenues[-1] > 0 and shares != 0:

        # -------- SENSITIVITY TABLE --------
        g_range = np.linspace(g_base - 0.02, g_base + 0.02, 5)
        w_range = np.linspace(wacc - 0.02, wacc + 0.02, 5)

        table = []

        for g in g_range:
            row = []
            for w in w_range:

                rev = revenues[-1]
                prev = rev
                fcf = []

                for t in range(years):
                    rev *= (1 + g)
                    delta = rev - prev

                    nopat = rev * em_base * (1 - tax_rate)
                    da_val = rev * da_base
                    capex_val = rev * cap_base
                    wc_val = delta * wc_base

                    fcf.append(nopat + da_val - capex_val - wc_val)
                    prev = rev

                if w <= terminal_growth:
                    row.append(0)
                    continue

                discounted = [fcf[i]/((1+w)**(i+1)) for i in range(len(fcf))]
                tv = fcf[-1]*(1+terminal_growth)/(w-terminal_growth)
                tv_discounted = tv/((1+w)**years)

                ev = sum(discounted) + tv_discounted
                equity = ev - net_debt

                row.append(round(equity/shares,2))

            table.append(row)

        df = pd.DataFrame(
            table,
            index=[round(g*100,1) for g in g_range],
            columns=[round(w*100,1) for w in w_range]
        )

        st.subheader("Sensitivity Table")
        st.caption("Rows = Growth (%) | Columns = WACC (%)")
        st.dataframe(df, use_container_width=True)

        # -------- TORNADO CHART --------
        st.subheader("Tornado Chart (Sensitivity Impact %)")

        def base_calc():
            rev = revenues[-1]
            prev = rev
            fcf = []

            for t in range(years):
                rev *= (1 + g_base)
                delta = rev - prev

                nopat = rev * em_base * (1 - tax_rate)
                da_val = rev * da_base
                capex_val = rev * cap_base
                wc_val = delta * wc_base

                fcf.append(nopat + da_val - capex_val - wc_val)
                prev = rev

            discounted = [fcf[i]/((1+wacc)**(i+1)) for i in range(len(fcf))]
            tv = fcf[-1]*(1+terminal_growth)/(wacc-terminal_growth)
            tv_discounted = tv/((1+wacc)**years)

            ev = sum(discounted) + tv_discounted
            equity = ev - net_debt

            return equity / shares

        base_value = base_calc()

        variables = {
            "Growth": g_base,
            "WACC": wacc,
            "EBIT Margin": em_base,
            "D&A": da_base,
            "Capex": cap_base,
            "WC": wc_base
        }

        impacts = {}

        for name, base in variables.items():

            low = base * 0.9
            high = base * 1.1

            def calc(val):
                rev = revenues[-1]
                prev = rev
                fcf = []

                for t in range(years):

                    g = val if name == "Growth" else g_base
                    w = val if name == "WACC" else wacc
                    em = val if name == "EBIT Margin" else em_base
                    da_p = val if name == "D&A" else da_base
                    cap_p = val if name == "Capex" else cap_base
                    wc_p = val if name == "WC" else wc_base

                    rev *= (1 + g)
                    delta = rev - prev

                    nopat = rev * em * (1 - tax_rate)
                    da_val = rev * da_p
                    capex_val = rev * cap_p
                    wc_val = delta * wc_p

                    fcf.append(nopat + da_val - capex_val - wc_val)
                    prev = rev

                if w <= terminal_growth:
                    return base_value

                discounted = [fcf[i]/((1+w)**(i+1)) for i in range(len(fcf))]
                tv = fcf[-1]*(1+terminal_growth)/(w-terminal_growth)
                tv_discounted = tv/((1+w)**years)

                ev = sum(discounted) + tv_discounted
                equity = ev - net_debt

                return equity / shares

            try:
                low_val = calc(low)
                high_val = calc(high)
                impacts[name] = ((high_val - low_val) / base_value) * 100
            except:
                impacts[name] = 0

        impacts = dict(sorted(impacts.items(), key=lambda x: abs(x[1])))

        names = list(impacts.keys())
        values = list(impacts.values())
        colors = ["green" if v > 0 else "red" for v in values]

        fig, ax = plt.subplots()
        ax.barh(names, values, color=colors)
        ax.axvline(0)
        ax.set_xlabel("% Impact on Valuation")

        for i, v in enumerate(values):
            ax.text(v, i, f"{round(v,1)}%", va='center')

        st.pyplot(fig)

    else:
        st.warning("Enter valid data to view sensitivity.")


# ---------------- SAVE FUNCTION ----------------
def save_inputs():
    output = BytesIO()

    fair_value = round(per_share,2) if per_share else ""
    mos_val = f"{round(((per_share-market_price)/market_price)*100,2)}%" if per_share and market_price else ""

    summary = pd.DataFrame({
        "Metric": ["Company", "Fair Value", "Market Price", "Margin of Safety"],
        "Value": [company_name, fair_value, market_price, mos_val]
    })

    assumptions = pd.DataFrame({
        "Field": ["Growth Rate","EBIT Margin","D&A %","Capex %","WC %","WACC","Terminal Growth","Tax Rate"],
        "Value": [
            growth_rate*100 if 'growth_rate' in globals() else "",
            ebit_margin*100 if 'ebit_margin' in globals() else "",
            da_pct_user*100 if 'da_pct_user' in globals() else "",
            capex_pct_user*100 if 'capex_pct_user' in globals() else "",
            wc_pct_user*100 if 'wc_pct_user' in globals() else "",
            wacc*100, terminal_growth*100, tax_rate*100
        ]
    })

    financials = pd.DataFrame({
        "Year": [start_year+i for i in range(5)],
        "Revenue": revenues,
        "EBIT": ebit,
        "D&A": da,
        "Capex": capex,
        "WC Change": wc
    })

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False, startrow=3)
        assumptions.to_excel(writer, sheet_name="Assumptions", index=False)
        financials.to_excel(writer, sheet_name="Financials", index=False)

        wb = writer.book

        ws = wb["Summary"]
        ws["A1"] = "DCF VALUATION REPORT"
        ws["A2"] = company_name
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"].font = Font(size=12)

    output.seek(0)
    return output

# -------- DOWNLOAD BUTTON --------
st.download_button(
    "💾 Download Full Report",
    data=save_inputs(),
    file_name=f"{company_name if company_name else 'DCF_Report'}.xlsx"
)


# ---------------- FOOTER ----------------
st.markdown("---")
st.markdown("<div style='text-align:center;color:gray'>Developed by Aqeel Ansari | +923334971073</div>", unsafe_allow_html=True)